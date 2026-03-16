from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from app.db.session import get_db
from app.models.estudiante import Estudiante, EstudianteEstado
from app.models.bootcamp import Bootcamp
from app.schemas.estudiante import EstudianteCreate, EstudianteUpdate, EstudianteSchema, EstudianteWithRelations, EstudianteKanban, BootcampSimple, UserSimple
from app.core.deps import require_student_success, get_current_user
from app.models.user import User
import openpyxl

router = APIRouter(prefix="/estudiantes", tags=["estudiantes"])


@router.get("", response_model=List[EstudianteSchema])
def get_estudiantes(
    skip: int = 0,
    limit: int = 100,
    bootcamp_id: Optional[int] = None,
    estado: Optional[EstudianteEstado] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    query = db.query(Estudiante)
    
    if bootcamp_id:
        query = query.filter(Estudiante.bootcamp_id == bootcamp_id)
    if estado:
        query = query.filter(Estudiante.estado == estado)
    
    estudiantes = query.offset(skip).limit(limit).all()
    
    result = []
    for est in estudiantes:
        bootcamp = db.query(Bootcamp).filter(Bootcamp.id == est.bootcamp_id).first()
        est_dict = {
            "nombre": est.nombre,
            "apellido": est.apellido,
            "email": est.email,
            "telefono": est.telefono,
            "whatsapp": est.whatsapp,
            "bootcamp_id": est.bootcamp_id,
            "estado": est.estado,
            "responsable_id": est.responsable_id,
            "id": est.id,
            "riesgo_desercion": est.riesgo_desercion,
            "fecha_ingreso": est.fecha_ingreso,
            "ultimo_contacto": est.ultimo_contacto,
            "ultimo_acceso_moodle": est.ultimo_acceso_moodle,
            "created_at": est.created_at,
            "bootcamp_nombre": bootcamp.nombre if bootcamp else None
        }
        result.append(est_dict)
    
    return result


@router.get("/kanban", response_model=dict)
def get_kanban(
    bootcamp_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    query = db.query(Estudiante)
    if bootcamp_id:
        query = query.filter(Estudiante.bootcamp_id == bootcamp_id)
    
    estudiantes = query.all()
    
    columns = {estado.value: [] for estado in EstudianteEstado}
    
    for est in estudiantes:
        bootcamp = db.query(Bootcamp).filter(Bootcamp.id == est.bootcamp_id).first()
        ultimo_contacto_dias = None
        if est.ultimo_contacto:
            ultimo_contacto = est.ultimo_contacto
            if ultimo_contacto.tzinfo is not None:
                ultimo_contacto = ultimo_contacto.replace(tzinfo=None)
            ultimo_contacto_dias = (datetime.now() - ultimo_contacto).days
        
        est_data = {
            "id": est.id,
            "nombre": est.nombre,
            "apellido": est.apellido,
            "email": est.email,
            "estado": est.estado.value,
            "riesgo_desercion": est.riesgo_desercion,
            "bootcamp_nombre": bootcamp.nombre if bootcamp else None,
            "ultimo_contacto_dias": ultimo_contacto_dias,
            "responsable_id": est.responsable_id
        }
        columns[est.estado.value].append(est_data)
    
    return columns


@router.get("/{estudiante_id}", response_model=EstudianteWithRelations)
def get_estudiante(
    estudiante_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    from app.models.nota import Nota
    from app.models.conversacion import Conversacion
    from app.models.ticket import Ticket
    
    estudiante = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante not found")
    
    result = EstudianteWithRelations.model_validate(estudiante)
    if estudiante.bootcamp:
        result.bootcamp = BootcampSimple(id=estudiante.bootcamp.id, nombre=estudiante.bootcamp.nombre)
    if estudiante.responsable:
        result.responsable = UserSimple(id=estudiante.responsable.id, nombre=estudiante.responsable.nombre, email=estudiante.responsable.email)
    
    notas = db.query(Nota).filter(Nota.estudiante_id == estudiante_id).order_by(Nota.fecha.desc()).limit(20).all()
    result.notas = [{"id": n.id, "contenido": n.contenido, "fecha": n.fecha, "autor": {"id": n.autor.id, "nombre": n.autor.nombre} if n.autor else None} for n in notas]
    
    conversaciones = db.query(Conversacion).filter(Conversacion.estudiante_id == estudiante_id).order_by(Conversacion.created_at.desc()).limit(20).all()
    result.conversaciones = [{"id": c.id, "tipo": c.tipo.value if hasattr(c.tipo, 'value') else c.tipo, "mensaje": c.mensaje, "created_at": c.created_at} for c in conversaciones]
    
    tickets = db.query(Ticket).filter(Ticket.estudiante_id == estudiante_id).order_by(Ticket.fecha_creacion.desc()).limit(20).all()
    result.tickets = [{"id": t.id, "tipo": t.tipo, "estado": t.estado.value, "fecha_creacion": t.fecha_creacion} for t in tickets]
    
    return result


@router.post("", response_model=EstudianteSchema)
def create_estudiante(
    estudiante_data: EstudianteCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    bootcamp = db.query(Bootcamp).filter(Bootcamp.id == estudiante_data.bootcamp_id).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp not found")
    
    existing = db.query(Estudiante).filter(Estudiante.email == estudiante_data.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    estudiante = Estudiante(**estudiante_data.model_dump())
    if not estudiante.responsable_id:
        estudiante.responsable_id = current_user.id
    
    db.add(estudiante)
    db.commit()
    db.refresh(estudiante)
    return estudiante


@router.patch("/{estudiante_id}", response_model=EstudianteSchema)
def update_estudiante(
    estudiante_id: int,
    estudiante_data: EstudianteUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    estudiante = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante not found")
    
    for key, value in estudiante_data.model_dump(exclude_unset=True).items():
        setattr(estudiante, key, value)
    
    db.commit()
    db.refresh(estudiante)
    return estudiante


@router.patch("/{estudiante_id}/status", response_model=EstudianteSchema)
def update_estudiante_status(
    estudiante_id: int,
    estado: EstudianteEstado,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    estudiante = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante not found")
    
    estudiante.estado = estado
    db.commit()
    db.refresh(estudiante)
    return estudiante


@router.delete("/{estudiante_id}")
def delete_estudiante(
    estudiante_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    estudiante = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante not found")
    
    db.delete(estudiante)
    db.commit()
    return {"message": "Estudiante deleted successfully"}


@router.post("/importar-excel")
def importar_estudiantes_excel(
    bootcamp_codigo: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    bootcamp = db.query(Bootcamp).filter(Bootcamp.codigo == bootcamp_codigo).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp no encontrado")
    
    wb = openpyxl.load_workbook(file.file)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    email_idx = None
    nombre_idx = None
    apellido_idx = None
    telefono_idx = None
    whatsapp_idx = None
    ultimo_acceso_idx = None
    
    for i, h in enumerate(headers):
        h_lower = str(h).lower() if h else ""
        if "email" in h_lower or "correo" in h_lower:
            email_idx = i
        elif "nombre" in h_lower and "apellido" not in h_lower:
            nombre_idx = i
        elif "apellido" in h_lower:
            apellido_idx = i
        elif "telefono" in h_lower:
            telefono_idx = i
        elif "whatsapp" in h_lower or "whats" in h_lower:
            whatsapp_idx = i
        elif "último acceso" in h_lower or "ultimo acceso" in h_lower or "last access" in h_lower:
            ultimo_acceso_idx = i
    
    if email_idx is None or nombre_idx is None:
        raise HTTPException(status_code=400, detail="El Excel debe tener columnas de email y nombre")
    
    creados = 0
    errores = []
    
    from datetime import datetime
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[email_idx]:
            continue
        
        email = str(row[email_idx]).strip().lower()
        nombre = str(row[nombre_idx]).strip() if nombre_idx and row[nombre_idx] else ""
        apellido = str(row[apellido_idx]).strip() if apellido_idx and row[apellido_idx] else ""
        telefono = str(row[telefono_idx]).strip() if telefono_idx and row[telefono_idx] else ""
        whatsapp = str(row[whatsapp_idx]).strip() if whatsapp_idx and row[whatsapp_idx] else ""
        
        ultimo_acceso_moodle = None
        if ultimo_acceso_idx and row[ultimo_acceso_idx]:
            try:
                valor = row[ultimo_acceso_idx]
                if isinstance(valor, datetime):
                    ultimo_acceso_moodle = valor
                elif isinstance(valor, str):
                    for fmt in ["%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"]:
                        try:
                            ultimo_acceso_moodle = datetime.strptime(valor.strip(), fmt)
                            break
                        except:
                            pass
            except:
                pass
        
        existing = db.query(Estudiante).filter(Estudiante.email == email).first()
        if existing:
            if ultimo_acceso_moodle:
                existing.ultimo_acceso_moodle = ultimo_acceso_moodle
            errores.append(f"{email}: actualizado")
            continue
        
        estudiante = Estudiante(
            email=email,
            nombre=nombre,
            apellido=apellido,
            telefono=telefono or None,
            whatsapp=whatsapp or None,
            bootcamp_id=bootcamp.id,
            estado=EstudianteEstado.NUEVO,
            responsable_id=current_user.id,
            ultimo_acceso_moodle=ultimo_acceso_moodle
        )
        db.add(estudiante)
        creados += 1
    
    db.commit()
    
    return {
        "message": f"Se importaron {creados} estudiantes",
        "errores": errores[:10],
        "bootcamp_id": bootcamp.id,
        "bootcamp_codigo": bootcamp.codigo
    }


@router.post("/evaluar-riesgo")
def evaluar_riesgo_y_crear_tickets(
    umbral_riesgo: int = 60,
    dias_sin_moodle: int = 7,
    dias_sin_contacto: int = 14,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    from app.models.ticket import Ticket, TicketTipo, TicketPrioridad
    
    estudiantes = db.query(Estudiante).all()
    tickets_creados = []
    estudiantes_evaluados = 0
    
    hoy = datetime.now(timezone.utc)
    
    for est in estudiantes:
        estudiantes_evaluados += 1
        
        # Calcular días sin acceso Moodle
        dias_moodle = None
        if est.ultimo_acceso_moodle:
            diff = hoy - est.ultimo_acceso_moodle
            dias_moodle = diff.days
        
        # Calcular días sin contacto
        dias_contacto = None
        if est.ultimo_contacto:
            diff = hoy - est.ultimo_contacto
            dias_contacto = diff.days
        
        # Calcular riesgo
        riesgo = 0
        
        # Por días sin Moodle (máximo 50 puntos)
        if dias_moodle is not None:
            if dias_moodle >= dias_sin_moodle:
                riesgo = max(riesgo, 50 + min(20, (dias_moodle - dias_sin_moodle) * 3))
        
        # Por días sin contacto (máximo 50 puntos)
        if dias_contacto is not None:
            if dias_contacto >= dias_sin_contacto:
                riesgo = max(riesgo, 50 + min(20, (dias_contacto - dias_sin_contacto) * 3))
        
        # Por estado de riesgo existente
        if est.estado == EstudianteEstado.EN_RIESGO:
            riesgo = max(riesgo, 70)
        elif est.estado == EstudianteEstado.NECESITA_SEGUIMIENTO:
            riesgo = max(riesgo, 40)
        
        # Actualizar riesgo en estudiante
        est.riesgo_desercion = min(riesgo, 100)
        
        # Actualizar estado según nivel de riesgo
        if riesgo >= 70:
            if est.estado != EstudianteEstado.EN_RIESGO and est.estado != EstudianteEstado.ABANDONO and est.estado != EstudianteEstado.GRADUADO:
                est.estado = EstudianteEstado.EN_RIESGO
        elif riesgo >= 40:
            if est.estado == EstudianteEstado.ACTIVO:
                est.estado = EstudianteEstado.NECESITA_SEGUIMIENTO
        elif riesgo < 30:
            # Si el riesgo bajo, volver a activo si estaba en riesgo o seguimiento
            if est.estado in [EstudianteEstado.EN_RIESGO, EstudianteEstado.NECESITA_SEGUIMIENTO]:
                est.estado = EstudianteEstado.ACTIVO
        
        # Si supera el umbral, crear ticket si no existe uno abierto
        if riesgo >= umbral_riesgo:
            # Verificar si ya existe un ticket abierto para este estudiante
            ticket_existente = db.query(Ticket).filter(
                Ticket.estudiante_id == est.id,
                Ticket.estado.in_(["abierto", "en_proceso", "espera"])
            ).first()
            
            if not ticket_existente:
                # Determinar prioridad según nivel de riesgo
                if riesgo >= 80:
                    prioridad = TicketPrioridad.URGENTE
                elif riesgo >= 70:
                    prioridad = TicketPrioridad.ALTA
                elif riesgo >= 60:
                    prioridad = TicketPrioridad.MEDIA
                else:
                    prioridad = TicketPrioridad.BAJA
                
                # Crear ticket
                ticket = Ticket(
                    estudiante_id=est.id,
                    tipo=TicketTipo.ASISTENCIA,
                    titulo=f"Alerta de riesgo de deserción - {est.nombre} {est.apellido}",
                    descripcion=f"Estudiante en riesgo de deserción ({riesgo}%)\n\n"
                              f"Días sin acceso Moodle: {dias_moodle if dias_moodle else 'N/A'}\n"
                              f"Días sin contacto: {dias_contacto if dias_contacto else 'N/A'}\n"
                              f"Estado actual: {est.estado.value}",
                    prioridad=prioridad,
                    asignado_a_id=est.responsable_id
                )
                db.add(ticket)
                tickets_creados.append({
                    "estudiante_id": est.id,
                    "nombre": f"{est.nombre} {est.apellido}",
                    "riesgo": riesgo,
                    "ticket_id": None  # Se asignará después del commit
                })
    
    db.commit()
    
    # Actualizar IDs de tickets
    for info in tickets_creados:
        est = db.query(Estudiante).filter(Estudiante.id == info["estudiante_id"]).first()
        if est:
            ticket = db.query(Ticket).filter(Ticket.estudiante_id == est.id).order_by(Ticket.id.desc()).first()
            if ticket:
                info["ticket_id"] = ticket.id
    
    return {
        "estudiantes_evaluados": estudiantes_evaluados,
        "tickets_creados": len(tickets_creados),
        "detalle": tickets_creados
    }
