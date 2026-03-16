from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime, timedelta
from app.db.session import get_db
from app.models.bootcamp import Bootcamp, BootcampEstado
from app.models.estudiante import Estudiante, EstudianteEstado
from app.models.nota import Nota
from app.models.conversacion import Conversacion, ConversacionTipo
from app.schemas.bootcamp import BootcampCreate, BootcampUpdate, BootcampSchema, BootcampWithCount
from app.core.deps import require_admin, require_student_success
from app.models.user import User
import openpyxl
import re
import unicodedata

def remove_accents(text):
    """Remove accents from text"""
    if not text:
        return text
    return ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')

router = APIRouter(prefix="/bootcamps", tags=["bootcamps"])


def parse_duration_to_datetime(duration_str: str) -> datetime | None:
    """Convierte duración como '2h 30m', '1 día', '3 días' a datetime restando de ahora"""
    if not duration_str:
        return None
    
    duration_str = str(duration_str).strip().lower()
    now = datetime.now()
    
    # Handle "Hoy" and "Ayer"
    if duration_str == "hoy":
        return now
    elif duration_str == "ayer":
        return now - timedelta(days=1)
    
    # Patrones de duración
    # "2h 30m", "3h 15m", "1h"
    horas_match = re.search(r'(\d+)\s*h', duration_str)
    minutos_match = re.search(r'(\d+)\s*m', duration_str)
    
    # "1 día", "2 días", "3 dia"
    dias_match = re.search(r'(\d+)\s*(?:día|dias|día)', duration_str)
    
    print(f"DEBUG parse: input='{duration_str}', horas_match={horas_match}, minutos_match={minutos_match}, dias_match={dias_match}")
    
    delta = timedelta()
    
    if horas_match:
        horas = int(horas_match.group(1))
        delta += timedelta(hours=horas)
    
    if minutos_match:
        minutos = int(minutos_match.group(1))
        delta += timedelta(minutes=minutos)
    
    if dias_match:
        dias = int(dias_match.group(1))
        delta += timedelta(days=dias)
    
    if delta.total_seconds() > 0:
        result = now - delta
        print(f"DEBUG parse result: {result}")
        return result
    
    return None


@router.get("", response_model=List[BootcampWithCount])
def get_bootcamps(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    bootcamps = db.query(Bootcamp).all()
    result = []
    for bootcamp in bootcamps:
        bootcamp_data = BootcampWithCount.model_validate(bootcamp)
        bootcamp_data.total_estudiantes = db.query(Estudiante).filter(Estudiante.bootcamp_id == bootcamp.id).count()
        bootcamp_data.estudiantes_en_riesgo = db.query(Estudiante).filter(
            Estudiante.bootcamp_id == bootcamp.id,
            Estudiante.riesgo_desercion >= 60
        ).count()
        bootcamp_data.estudiantes_activos = max(
            bootcamp_data.total_estudiantes - bootcamp_data.estudiantes_en_riesgo,
            0
        )
        result.append(bootcamp_data)
    return result


@router.get("/{bootcamp_id}", response_model=BootcampSchema)
def get_bootcamp(
    bootcamp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    bootcamp = db.query(Bootcamp).filter(Bootcamp.id == bootcamp_id).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp not found")
    return bootcamp


@router.post("", response_model=BootcampSchema)
def create_bootcamp(
    bootcamp_data: BootcampCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    bootcamp = Bootcamp(**bootcamp_data.model_dump())
    db.add(bootcamp)
    db.commit()
    db.refresh(bootcamp)
    return bootcamp


@router.put("/{bootcamp_id}", response_model=BootcampSchema)
def update_bootcamp(
    bootcamp_id: int,
    bootcamp_data: BootcampUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    bootcamp = db.query(Bootcamp).filter(Bootcamp.id == bootcamp_id).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp not found")
    
    for key, value in bootcamp_data.model_dump(exclude_unset=True).items():
        setattr(bootcamp, key, value)
    
    db.commit()
    db.refresh(bootcamp)
    return bootcamp


@router.delete("/{bootcamp_id}")
def delete_bootcamp(
    bootcamp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    bootcamp = db.query(Bootcamp).filter(Bootcamp.id == bootcamp_id).first()
    if not bootcamp:
        raise HTTPException(status_code=404, detail="Bootcamp not found")
    
    # Primero borrar los estudiantes relacionados
    db.query(Estudiante).filter(Estudiante.bootcamp_id == bootcamp_id).delete()
    
    # Luego borrar el bootcamp
    db.delete(bootcamp)
    db.commit()
    return {"message": "Bootcamp deleted successfully"}


@router.post("/importar-excel")
def importar_bootcamp_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_student_success)
):
    wb = openpyxl.load_workbook(file.file)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    codigo_idx = None
    email_idx = None
    nombre_est_idx = None
    apellido_idx = None
    telefono_idx = None
    whatsapp_idx = None
    ultimo_acceso_idx = None
    ultimo_contacto_idx = None
    notas_idx = None
    conversaciones_idx = None
    
    for i, h in enumerate(headers):
        h_lower = str(h).lower() if h else ""
        h_stripped = h_lower.strip() if h else ""
        h_normalized = remove_accents(h_stripped)
        
        # Código bootcamp - busca RTD- o cualquier cosa que parezca código
        if codigo_idx is None:
            if "rtd-" in h_lower or "codigo" in h_normalized or "codigo" in h_lower:
                codigo_idx = i
            # Si es la primera columna y tiene formato de código
            elif i == 0 and h and ("-" in str(h) or len(str(h).strip()) > 5):
                codigo_idx = i
        
        # Nombre del estudiante - "Nombre" sin "bootcamp"
        if nombre_est_idx is None:
            if h_stripped == "nombre":
                nombre_est_idx = i
            elif "nombre" in h_lower and "bootcamp" not in h_lower and "apellido" not in h_lower:
                nombre_est_idx = i
        
        # Email - "Dirección de correo" o "Email"
        if email_idx is None:
            if "correo" in h_lower or "email" in h_lower:
                email_idx = i
        
        # Apellido - "Apellido(s)"
        if apellido_idx is None:
            if "apellido" in h_lower:
                apellido_idx = i
        
        # Teléfono - "Teléfono"
        if telefono_idx is None:
            if "telefono" in h_lower or "teléfono" in h_lower:
                telefono_idx = i
        
        # WhatsApp - "WhatsApp"
        if whatsapp_idx is None:
            if "whatsapp" in h_lower or "whats" in h_lower:
                whatsapp_idx = i
        
        # Último acceso - "Último acceso al curso"
        if ultimo_acceso_idx is None:
            if "último acceso" in h_lower or "ultimo acceso" in h_lower or "last access" in h_lower:
                ultimo_acceso_idx = i
        
        # Último contacto - "Último contacto"
        if ultimo_contacto_idx is None:
            if "último contacto" in h_lower or "ultimo contacto" in h_lower or "last contact" in h_lower:
                ultimo_contacto_idx = i
        
        # Notas de seguimiento - "Notas de seguimiento"
        if notas_idx is None:
            if "nota" in h_lower and "seguimiento" in h_lower:
                notas_idx = i
        
        # Conversaciones - "Conversaciones"
        if conversaciones_idx is None:
            if "conversacion" in h_lower or "conversación" in h_lower:
                conversaciones_idx = i
    
    # Debug: mostrar qué columnas se detectaron
    print(f"DEBUG - Columnas detectadas:")
    print(f"  codigo_idx: {codigo_idx} -> {headers[codigo_idx] if codigo_idx is not None else 'NO ENCONTRADO'}")
    print(f"  nombre_est_idx: {nombre_est_idx} -> {headers[nombre_est_idx] if nombre_est_idx else 'NO ENCONTRADO'}")
    print(f"  email_idx: {email_idx} -> {headers[email_idx] if email_idx else 'NO ENCONTRADO'}")
    print(f"  apellido_idx: {apellido_idx} -> {headers[apellido_idx] if apellido_idx else 'NO ENCONTRADO'}")
    print(f"  telefono_idx: {telefono_idx} -> {headers[telefono_idx] if telefono_idx else 'NO ENCONTRADO'}")
    print(f"  whatsapp_idx: {whatsapp_idx} -> {headers[whatsapp_idx] if whatsapp_idx else 'NO ENCONTRADO'}")
    print(f"  ultimo_acceso_idx: {ultimo_acceso_idx} -> {headers[ultimo_acceso_idx] if ultimo_acceso_idx else 'NO ENCONTRADO'}")
    print(f"  ultimo_contacto_idx: {ultimo_contacto_idx} -> {headers[ultimo_contacto_idx] if ultimo_contacto_idx else 'NO ENCONTRADO'}")
    print(f"  notas_idx: {notas_idx} -> {headers[notas_idx] if notas_idx else 'NO ENCONTRADO'}")
    print(f"  conversaciones_idx: {conversaciones_idx} -> {headers[conversaciones_idx] if conversaciones_idx else 'NO ENCONTRADO'}")
    
    if codigo_idx is None:
        raise HTTPException(status_code=400, detail=f"No se encontró el código del bootcamp. Headers: {headers}")
    
    if email_idx is None:
        return {
            "message": "No se encontró columna de email",
            "estudiantes_creados": 0
        }
    
    if nombre_est_idx is None:
        raise HTTPException(status_code=400, detail="No se encontró columna de nombre del estudiante")
    
    creados = 0
    errores = []
    
    notas_creadas = 0
    conversaciones_creadas = 0
    
    # Cache de bootcamps para evitar consultas repetidas
    bootcamps_cache = {}
    current_bootcamp = None
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[email_idx]:
            continue
        
        try:
            # Obtener código del bootcamp de esta fila
            codigo_fila = None
            if codigo_idx is not None and row[codigo_idx]:
                codigo_fila = str(row[codigo_idx]).strip()
            
            if not codigo_fila:
                email_val = row[email_idx] if email_idx and row[email_idx] else '?'
                errores.append(f"{email_val}: código de bootcamp vacío")
                continue
            
            # Obtener o crear el bootcamp para esta fila
            if codigo_fila not in bootcamps_cache:
                bootcamp = db.query(Bootcamp).filter(Bootcamp.codigo == codigo_fila).first()
                if not bootcamp:
                    bootcamp = Bootcamp(
                        codigo=codigo_fila,
                        nombre=codigo_fila,
                        estado=BootcampEstado.ACTIVO
                    )
                    db.add(bootcamp)
                    db.commit()
                    db.refresh(bootcamp)
                bootcamps_cache[codigo_fila] = bootcamp
            
            current_bootcamp = bootcamps_cache[codigo_fila]
            
            email = str(row[email_idx]).strip().lower()
            nombre = str(row[nombre_est_idx]).strip() if nombre_est_idx and row[nombre_est_idx] else ""
            apellido = str(row[apellido_idx]).strip() if apellido_idx and row[apellido_idx] else ""
            telefono = str(row[telefono_idx]).strip() if telefono_idx and row[telefono_idx] else ""
            whatsapp = str(row[whatsapp_idx]).strip() if whatsapp_idx and row[whatsapp_idx] else ""
            
            # Parsear último acceso (puede ser fecha o duración)
            ultimo_acceso_moodle = None
            if ultimo_acceso_idx is not None and row[ultimo_acceso_idx]:
                try:
                    valor = row[ultimo_acceso_idx]
                    print(f"DEBUG ultimo_acceso: index={ultimo_acceso_idx}, valor='{valor}', type={type(valor)}")
                    if isinstance(valor, datetime):
                        ultimo_acceso_moodle = valor
                    elif isinstance(valor, str):
                        # Primero intentar parsing de fecha
                        parsed = False
                        for fmt in ["%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"]:
                            try:
                                ultimo_acceso_moodle = datetime.strptime(valor.strip(), fmt)
                                parsed = True
                                break
                            except:
                                pass
                        # Si no es fecha, intentar parsear duración
                        if not parsed:
                            ultimo_acceso_moodle = parse_duration_to_datetime(valor)
                except:
                    pass
            
            # Parsear último contacto (duración)
            ultimo_contacto = None
            if ultimo_contacto_idx is not None and row[ultimo_contacto_idx]:
                try:
                    valor = row[ultimo_contacto_idx]
                    print(f"DEBUG ultimo_contacto: index={ultimo_contacto_idx}, valor='{valor}', type={type(valor)}")
                    if isinstance(valor, str):
                        ultimo_contacto = parse_duration_to_datetime(valor)
                except:
                    pass
            
            # Obtener notas y conversaciones
            nota_contenido = None
            if notas_idx and row[notas_idx]:
                nota_contenido = str(row[notas_idx]).strip() if row[notas_idx] else None
            
            conversacion_mensaje = None
            if conversaciones_idx and row[conversaciones_idx]:
                conversacion_mensaje = str(row[conversaciones_idx]).strip() if row[conversaciones_idx] else None
            
            existing = db.query(Estudiante).filter(Estudiante.email == email).first()
            if existing:
                # Actualizar campos si existen
                if ultimo_acceso_moodle:
                    existing.ultimo_acceso_moodle = ultimo_acceso_moodle
                if ultimo_contacto:
                    existing.ultimo_contacto = ultimo_contacto
                errores.append(f"{email}: ya existe")
                continue
            
            estudiante = Estudiante(
                email=email,
                nombre=nombre,
                apellido=apellido,
                telefono=telefono or None,
                whatsapp=whatsapp or None,
                bootcamp_id=current_bootcamp.id,
                estado=EstudianteEstado.NUEVO,
                responsable_id=current_user.id,
                ultimo_acceso_moodle=ultimo_acceso_moodle,
                ultimo_contacto=ultimo_contacto
            )
            db.add(estudiante)
            db.flush()  # Para obtener el ID del estudiante
            
            # Crear nota de seguimiento si existe
            if nota_contenido:
                nota = Nota(
                    estudiante_id=estudiante.id,
                    autor_id=current_user.id,
                    contenido=nota_contenido
                )
                db.add(nota)
                notas_creadas += 1
            
            # Crear conversación si existe
            if conversacion_mensaje:
                conversacion = Conversacion(
                    estudiante_id=estudiante.id,
                    tipo=ConversacionTipo.EMAIL,  # Default como email
                    mensaje=conversacion_mensaje,
                    usuario_id=current_user.id
                )
                db.add(conversacion)
                conversaciones_creadas += 1
            
            creados += 1
        except Exception as e:
            errores.append(f"Error en fila: {str(e)}")
    
    db.commit()
    
    bootcamps_resumen = []
    for codigo, bc in bootcamps_cache.items():
        count = db.query(Estudiante).filter(Estudiante.bootcamp_id == bc.id).count()
        bootcamps_resumen.append({
            "codigo": bc.codigo,
            "nombre": bc.nombre,
            "estudiantes": count
        })
    
    return {
        "message": f"Se importaron {creados} estudiantes en {len(bootcamps_resumen)} bootcamps",
        "bootcamps": bootcamps_resumen,
        "estudiantes_creados": creados,
        "bootcamps_creados": len(bootcamps_resumen),
        "notas_creadas": notas_creadas,
        "conversaciones_creadas": conversaciones_creadas,
        "errores": errores[:10]
    }
